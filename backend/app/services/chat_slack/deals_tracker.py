"""Deals Tracker diff for Todd: which deals are newly up for discussion.

Every week someone drops a `Deals Tracker <date>.xlsx` into the
#existing_pipeline Slack channel listing the pipeline deals for the
upcoming meeting. This module compares the latest tracker (or the one
for a given week) against the prior week's and returns the deals that
are *newly* "to be discussed".

Definitions (confirmed with Rom 2026-06-16):
  * "to be discussed" = any Status other than 'Warming Station'.
  * "new"             = to-be-discussed in the latest tracker AND not
                        to-be-discussed in the previous one -- i.e. it
                        was either absent last week, or present only as
                        'Warming Station'.

Each tracker row's "Deal Name" is the deal codename (e.g. 'Project Auto
II'), which matches dealcloud.deal.name -- so a returned deal_name can be
fed straight into get_deal_one_pager.

Slack access: the files live in a channel the *bot* token can't read
(no files:read scope, not a member), so this uses SLACK_USER_TOKEN.
Files are listed via slack_sdk, downloaded with stdlib urllib (no extra
HTTP dep), and parsed with openpyxl.
"""
from __future__ import annotations

import io
import urllib.request
from datetime import date, datetime, timezone
from typing import Optional

from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError

from ...config import settings

# #existing_pipeline -- the weekly "Deals Tracker <date>.xlsx" drop.
PIPELINE_CHANNEL_ID = "CECA2KU49"
PIPELINE_CHANNEL_NAME = "#existing_pipeline"
TRACKER_NAME_SUBSTR = "deals tracker"          # matched case-insensitively
EXCLUDED_STATUS = "warming station"            # NOT "to be discussed"


class TrackerError(Exception):
    """Recoverable problem (misconfig / missing data) surfaced to the
    model as a friendly message rather than a 500."""


# ---------------------------------------------------------------------------
# Slack file access
# ---------------------------------------------------------------------------

def _client() -> WebClient:
    token = settings.slack_user_token
    if not token:
        raise TrackerError(
            "SLACK_USER_TOKEN is not configured, so I can't read the "
            "Deals Tracker files from Slack."
        )
    return WebClient(token=token)


def _list_tracker_files(client: WebClient) -> list[dict]:
    """All 'Deals Tracker' spreadsheets in #existing_pipeline, newest
    first. Each item: {name, created (epoch int), url}."""
    out: list[dict] = []
    page = 1
    while True:
        try:
            resp = client.files_list(
                channel=PIPELINE_CHANNEL_ID, types="spreadsheets",
                count=200, page=page,
            )
        except SlackApiError as e:
            raise TrackerError(
                f"Slack rejected the file listing ({e.response.get('error')}). "
                "The SLACK_USER_TOKEN may lack files:read or channel access."
            ) from e
        for f in resp.get("files", []):
            if TRACKER_NAME_SUBSTR in (f.get("name") or "").lower():
                url = f.get("url_private_download") or f.get("url_private")
                if url:
                    out.append({
                        "name": f["name"],
                        "created": int(f.get("created", 0)),
                        "url": url,
                    })
        paging = resp.get("paging", {}) or {}
        if page >= paging.get("pages", 1):
            break
        page += 1
    out.sort(key=lambda f: f["created"], reverse=True)
    return out


def _download(url: str) -> bytes:
    req = urllib.request.Request(
        url, headers={"Authorization": f"Bearer {settings.slack_user_token}"}
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
    except Exception as e:  # noqa: BLE001 -- surface any network error
        raise TrackerError(f"Couldn't download a tracker file from Slack: {e}") from e
    # Slack returns an HTML login page (not the file) when auth fails.
    if data[:4] != b"PK\x03\x04":
        raise TrackerError(
            "Slack returned something that isn't an .xlsx file (likely an "
            "auth problem with SLACK_USER_TOKEN)."
        )
    return data


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _norm(name: str) -> str:
    """Match key for a deal name across two trackers: trimmed, lowered,
    internal whitespace collapsed."""
    return " ".join(str(name).split()).lower()


def _parse_tracker(content: bytes) -> dict[str, tuple[str, str]]:
    """Parse one tracker workbook into {norm_name: (orig_name, status)}.

    Finds the header row (the one carrying both 'Deal Name' and 'Status')
    rather than assuming row 0, so a future layout tweak doesn't silently
    break. Raises TrackerError if those columns can't be found.
    """
    from openpyxl import load_workbook  # lazy: keep tools import cheap

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    name_col = status_col = None
    deals: dict[str, tuple[str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = ["" if c is None else str(c).strip() for c in row]
        if name_col is None:
            lowered = [c.lower() for c in cells]
            if "deal name" in lowered and "status" in lowered:
                name_col = lowered.index("deal name")
                status_col = lowered.index("status")
            if i > 15:  # header should be near the top; give up scanning
                break
            continue
        if name_col >= len(cells) or status_col >= len(cells):
            continue
        deal_name = cells[name_col]
        if not deal_name:
            continue
        deals[_norm(deal_name)] = (deal_name, cells[status_col])

    if name_col is None:
        raise TrackerError(
            "Couldn't find 'Deal Name' and 'Status' columns in the tracker "
            "spreadsheet -- its layout may have changed."
        )
    return deals


def _discussable(status: str) -> bool:
    return _norm(status) != EXCLUDED_STATUS


def _posted_date(created: int) -> datetime:
    return datetime.fromtimestamp(created, tz=timezone.utc)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def compute_new_deals_to_discuss(as_of_date: Optional[date] = None) -> dict:
    """Diff the relevant pair of Deals Tracker files and return the deals
    newly up for discussion.

    as_of_date: if given, the "new" tracker is the latest one posted on or
    before that date, compared to the one before it. If None, the two most
    recent trackers are used.
    """
    client = _client()
    files = _list_tracker_files(client)
    if len(files) < 2:
        raise TrackerError(
            f"Found {len(files)} 'Deals Tracker' file(s) in "
            f"{PIPELINE_CHANNEL_NAME}; need at least two to compare."
        )

    if as_of_date is None:
        new_idx = 0
    else:
        new_idx = next(
            (i for i, f in enumerate(files)
             if _posted_date(f["created"]).date() <= as_of_date),
            None,
        )
        if new_idx is None:
            raise TrackerError(
                f"No 'Deals Tracker' was posted on or before "
                f"{as_of_date.isoformat()} in {PIPELINE_CHANNEL_NAME}."
            )
        if new_idx + 1 >= len(files):
            raise TrackerError(
                f"The tracker for {as_of_date.isoformat()} is the earliest "
                "one available, so there's nothing before it to compare to."
            )

    new_file = files[new_idx]
    prev_file = files[new_idx + 1]

    new_map = _parse_tracker(_download(new_file["url"]))
    prev_map = _parse_tracker(_download(prev_file["url"]))

    new_deals: list[dict] = []
    for key, (orig_name, status) in sorted(new_map.items(), key=lambda kv: kv[1][0].lower()):
        if not _discussable(status):
            continue
        prev = prev_map.get(key)
        if prev is not None and _discussable(prev[1]):
            continue  # already up for discussion last week -- not new
        new_deals.append({
            "deal_name": orig_name,
            "status": status,
            "previous_status": prev[1] if prev else None,
            "change": ("absent from last week's tracker" if prev is None
                       else f"was '{prev[1]}' last week"),
        })

    def _file_info(f: dict, deal_map: dict) -> dict:
        dt = _posted_date(f["created"])
        return {
            "file": f["name"],
            "posted_date": dt.date().isoformat(),
            "posted_at": dt.isoformat(),
            "deal_count": len(deal_map),
        }

    return {
        "channel": PIPELINE_CHANNEL_NAME,
        "new_tracker": _file_info(new_file, new_map),
        "previous_tracker": _file_info(prev_file, prev_map),
        "criteria": (
            "A deal is 'new to discuss' if its status is not 'Warming "
            "Station' in the new tracker AND it was not already up for "
            "discussion in the previous one (absent, or only 'Warming "
            "Station' before)."
        ),
        "new_deals_count": len(new_deals),
        "new_deals": new_deals,
        "present_instructions": (
            "List these deals with their status. Each deal_name matches a "
            "deal codename -- if the user wants the one-pagers, call "
            "get_deal_one_pager(deal_name) for each. If new_deals is empty, "
            "say no new deals came up for discussion versus last week."
        ),
    }
